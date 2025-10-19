from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse, Http404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, Min, Max
from django.db.models.functions import Extract
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import TaxInvoice, TaxInvoiceItem
from orders.models import Order
from .utils import generate_invoice_pdf
from django.forms import modelformset_factory


def is_staff_or_superuser(user):
    return user.is_staff or user.is_superuser


@user_passes_test(is_staff_or_superuser)
def invoice_list(request):
    """Display list of invoices with date filtering"""
    invoices = TaxInvoice.objects.all().select_related('order', 'order__user')
    
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            invoices = invoices.filter(invoice_date__date__gte=start_date)
        except ValueError:
            messages.error(request, 'Invalid start date format')
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            invoices = invoices.filter(invoice_date__date__lte=end_date)
        except ValueError:
            messages.error(request, 'Invalid end date format')
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(customer_name__icontains=search) |
            Q(customer_email__icontains=search) |
            Q(order__order_number__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(invoices, 25)  # 25 invoices per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'invoices': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'search': search,
        'total_count': invoices.count()
    }
    
    return render(request, 'invoices/invoice_list.html', context)


@user_passes_test(is_staff_or_superuser)
def generate_invoice(request, order_id):
    """Generate tax invoice for an order"""
    order = get_object_or_404(Order, id=order_id)
    
    # Check if invoice already exists
    if hasattr(order, 'tax_invoice'):
        messages.info(request, f'Invoice already exists for order {order.order_number}')
        return redirect('invoices:invoice_detail', invoice_id=order.tax_invoice.id)
    
    try:
        # Create tax invoice
        invoice = TaxInvoice.objects.create(order=order)
        invoice.populate_from_order()
        
        # Create invoice items from order items
        for order_item in order.items.all():
            # Get product-specific tax rate
            product_tax_rate = getattr(order_item.product, 'tax_percentage', 5.00)
            
            # Note: order_item.price is tax-inclusive, TaxInvoiceItem will separate tax in save()
            TaxInvoiceItem.objects.create(
                invoice=invoice,
                product_name=order_item.product.name,
                product_description=order_item.product.description[:200],
                hsn_code=getattr(order_item.product, 'hsn_code', '') or '',
                size=order_item.size or '',
                quantity=order_item.quantity,
                unit_price=order_item.price,  # Tax-inclusive price from order
                tax_percentage=product_tax_rate,
            )
        
        # Generate PDF
        pdf_generated = generate_invoice_pdf(invoice)
        
        if pdf_generated:
            invoice.is_generated = True
            invoice.save()
            messages.success(request, f'Invoice {invoice.invoice_number} generated successfully')
        else:
            messages.warning(request, f'Invoice {invoice.invoice_number} created but PDF generation failed')
        
        return redirect('invoices:invoice_detail', invoice_id=invoice.id)
        
    except Exception as e:
        messages.error(request, f'Error generating invoice: {str(e)}')
        return redirect('admin:orders_order_change', order_id)


@user_passes_test(is_staff_or_superuser)
def invoice_detail(request, invoice_id):
    """Display invoice details"""
    invoice = get_object_or_404(TaxInvoice, id=invoice_id)
    
    # Check if this is a print request
    is_print = request.GET.get('print', '').lower() == 'true'
    
    context = {
        'invoice': invoice,
        'invoice_items': invoice.items.all(),
    }
    
    # Use clean print-only template if requested
    if is_print:
        return render(request, 'invoices/invoice_print_clean.html', context)
    
    return render(request, 'invoices/invoice_detail.html', context)


@user_passes_test(is_staff_or_superuser)
def invoice_pdf(request, invoice_id):
    """Serve invoice PDF. Regenerate if missing or if invoice has no items. Support ?download=1 and ?regen=1."""
    invoice = get_object_or_404(TaxInvoice, id=invoice_id)
    
    force_regen = request.GET.get('regen') in ('1', 'true', 'yes')
    if force_regen or (not invoice.invoice_pdf) or (not invoice.items.exists()):
        pdf_generated = generate_invoice_pdf(invoice)
        if not pdf_generated:
            raise Http404('Invoice PDF not found and could not be generated')
    
    try:
        response = HttpResponse(invoice.invoice_pdf.read(), content_type='application/pdf')
        disposition = 'attachment' if request.GET.get('download') in ('1','true','yes') else 'inline'
        response['Content-Disposition'] = f'{disposition}; filename="{invoice.invoice_number}.pdf"'
        return response
    except Exception as e:
        raise Http404(f'Error serving PDF: {str(e)}')


@user_passes_test(is_staff_or_superuser)
def export_invoices_excel(request):
    """Export invoices to Excel"""
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    invoices = TaxInvoice.objects.all().select_related('order', 'order__user')
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            invoices = invoices.filter(invoice_date__date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            invoices = invoices.filter(invoice_date__date__lte=end_date)
        except ValueError:
            pass
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tax Invoices'
    
    # Headers
    headers = [
        'Invoice Number', 'Order Number', 'Invoice Date', 'Customer Name', 
        'Customer Email', 'Shipping City', 'Shipping State', 
        'Subtotal', 'CGST Amount', 'SGST Amount', 'IGST Amount', 
        'Total Tax', 'Discount', 'Final Amount', 'Status'
    ]
    
    # Style headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Auto-adjust column width
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = max(len(header) + 2, 12)
    
    # Add data
    for row_num, invoice in enumerate(invoices, 2):
        ws.cell(row=row_num, column=1, value=invoice.invoice_number)
        ws.cell(row=row_num, column=2, value=invoice.order.order_number)
        ws.cell(row=row_num, column=3, value=invoice.invoice_date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_num, column=4, value=invoice.customer_name)
        ws.cell(row=row_num, column=5, value=invoice.customer_email)
        ws.cell(row=row_num, column=6, value=invoice.shipping_city)
        ws.cell(row=row_num, column=7, value=invoice.shipping_state)
        ws.cell(row=row_num, column=8, value=float(invoice.subtotal))
        ws.cell(row=row_num, column=9, value=float(invoice.cgst_amount))
        ws.cell(row=row_num, column=10, value=float(invoice.sgst_amount))
        ws.cell(row=row_num, column=11, value=float(invoice.igst_amount))
        ws.cell(row=row_num, column=12, value=float(invoice.total_tax))
        ws.cell(row=row_num, column=13, value=float(invoice.discount_amount))
        ws.cell(row=row_num, column=14, value=float(invoice.final_amount))
        ws.cell(row=row_num, column=15, value='Generated' if invoice.is_generated else 'Pending')
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    from django.utils import timezone
    now_local = timezone.localtime(timezone.now())
    filename = f'tax_invoices_{now_local.strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# =============================================================================
# ADMIN DASHBOARD AND MANAGEMENT VIEWS
# =============================================================================

@user_passes_test(is_staff_or_superuser)
def admin_dashboard(request):
    """Admin dashboard for invoice management"""
    # Get statistics
    total_invoices = TaxInvoice.objects.count()
    generated_invoices = TaxInvoice.objects.filter(is_generated=True).count()
    pending_invoices = total_invoices - generated_invoices
    
    # Monthly statistics
    current_month = datetime.now().replace(day=1)
    monthly_invoices = TaxInvoice.objects.filter(created_at__gte=current_month).count()
    monthly_revenue = TaxInvoice.objects.filter(
        created_at__gte=current_month
    ).aggregate(total=Sum('final_amount'))['total'] or 0
    
    # Recent invoices
    recent_invoices = TaxInvoice.objects.select_related('order', 'order__user').order_by('-created_at')[:5]
    
    # Orders without invoices
    orders_without_invoices = Order.objects.filter(
        status__in=['confirmed', 'shipped']
    ).exclude(tax_invoice__isnull=False)[:5]
    
    context = {
        'total_invoices': total_invoices,
        'generated_invoices': generated_invoices,
        'pending_invoices': pending_invoices,
        'monthly_invoices': monthly_invoices,
        'monthly_revenue': monthly_revenue,
        'recent_invoices': recent_invoices,
        'orders_without_invoices': orders_without_invoices,
    }
    
    return render(request, 'invoices/admin_dashboard.html', context)


@user_passes_test(is_staff_or_superuser)
def reports_dashboard(request):
    """Reports dashboard for invoice analytics"""
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    invoices = TaxInvoice.objects.all()
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            invoices = invoices.filter(invoice_date__date__gte=start_date)
        except ValueError:
            start_date = None
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            invoices = invoices.filter(invoice_date__date__lte=end_date)
        except ValueError:
            end_date = None
    
    # Analytics
    total_revenue = invoices.aggregate(total=Sum('final_amount'))['total'] or 0
    total_tax = invoices.aggregate(total=Sum('total_tax'))['total'] or 0
    total_discounts = invoices.aggregate(total=Sum('discount_amount'))['total'] or 0
    avg_invoice_value = invoices.aggregate(avg=Sum('final_amount'))['avg'] or 0
    
    # Top customers by revenue
    top_customers = invoices.values('customer_name', 'customer_email').annotate(
        total_spent=Sum('final_amount'),
        invoice_count=Count('id')
    ).order_by('-total_spent')[:10]
    
    # Monthly breakdown
    monthly_data = invoices.extra({
        'month': 'EXTRACT(month FROM invoice_date)',
        'year': 'EXTRACT(year FROM invoice_date)'
    }).values('month', 'year').annotate(
        count=Count('id'),
        revenue=Sum('final_amount')
    ).order_by('-year', '-month')[:12]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_invoices': invoices.count(),
        'total_revenue': total_revenue,
        'total_tax': total_tax,
        'total_discounts': total_discounts,
        'avg_invoice_value': avg_invoice_value / invoices.count() if invoices.count() > 0 else 0,
        'top_customers': top_customers,
        'monthly_data': monthly_data,
    }
    
    return render(request, 'invoices/reports_dashboard.html', context)


# =============================================================================
# CUSTOMER-FACING VIEWS
# =============================================================================

@login_required
def customer_invoices(request):
    """Customer can view their own invoices"""
    user_orders = Order.objects.filter(user=request.user)
    invoices = TaxInvoice.objects.filter(order__in=user_orders).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(invoices, 10)  # 10 invoices per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'invoices': page_obj,
        'user': request.user,
    }
    
    return render(request, 'invoices/customer_invoices.html', context)


@login_required
def customer_invoice_detail(request, invoice_id):
    """Customer can view their own invoice details"""
    invoice = get_object_or_404(
        TaxInvoice,
        id=invoice_id,
        order__user=request.user  # Ensure user can only see their own invoices
    )
    
    context = {
        'invoice': invoice,
        'invoice_items': invoice.items.all(),
        'user': request.user,
    }
    
    return render(request, 'invoices/customer_invoice_detail.html', context)


@login_required
def customer_invoice_pdf(request, invoice_id):
    """Customer can download their own invoice PDF. Regenerate if missing or if invoice has no items. Support ?regen=1"""
    invoice = get_object_or_404(
        TaxInvoice,
        id=invoice_id,
        order__user=request.user  # Ensure user can only access their own invoices
    )
    
    force_regen = request.GET.get('regen') in ('1','true','yes')
    if force_regen or (not invoice.invoice_pdf) or (not invoice.items.exists()):
        pdf_generated = generate_invoice_pdf(invoice)
        if not pdf_generated:
            raise Http404('Invoice PDF not found and could not be generated')
    
    try:
        response = HttpResponse(invoice.invoice_pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Invoice-{invoice.invoice_number}.pdf"'
        return response
    except Exception as e:
        raise Http404(f'Error serving PDF: {str(e)}')


# =============================================================================
# PRODUCT AMOUNT MANAGEMENT VIEWS
# =============================================================================

@user_passes_test(is_staff_or_superuser)
def manage_invoice_items(request, invoice_id):
    """Manage product amounts for a specific invoice"""
    invoice = get_object_or_404(TaxInvoice, id=invoice_id)
    
    # Create a formset for editing invoice items
    TaxInvoiceItemFormSet = modelformset_factory(
        TaxInvoiceItem,
        fields=['product_name', 'quantity', 'unit_price', 'tax_percentage', 'hsn_code', 'size'],
        extra=1,  # Allow adding one new item
        can_delete=True
    )
    
    if request.method == 'POST':
        formset = TaxInvoiceItemFormSet(
            request.POST,
            queryset=TaxInvoiceItem.objects.filter(invoice=invoice)
        )
        
        if formset.is_valid():
            instances = formset.save(commit=False)
            
            # Set the invoice for new items and save
            for instance in instances:
                instance.invoice = invoice
                # Tax calculation is handled in the model's save method
                instance.save()
            
            # Delete marked items
            for obj in formset.deleted_objects:
                obj.delete()
            
            # Recalculate invoice totals
            invoice.subtotal = sum(item.total_price for item in invoice.items.all())
            invoice.calculate_taxes()
            invoice.save()
            
            # Regenerate PDF if it exists
            if invoice.invoice_pdf:
                generate_invoice_pdf(invoice)
            
            messages.success(request, f'Invoice {invoice.invoice_number} items updated successfully')
            return redirect('invoices:admin_invoice_detail', invoice_id=invoice.id)
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        formset = TaxInvoiceItemFormSet(
            queryset=TaxInvoiceItem.objects.filter(invoice=invoice)
        )
    
    context = {
        'invoice': invoice,
        'formset': formset,
    }
    
    return render(request, 'invoices/manage_invoice_items.html', context)


# =============================================================================
# ORDER EXPORT FUNCTIONALITY
# =============================================================================

@user_passes_test(is_staff_or_superuser)
def export_orders_excel(request):
    """Export detailed orders report to Excel"""
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status_filter = request.GET.get('status')
    
    orders = Order.objects.all().select_related('user', 'coupon').prefetch_related('items__product')
    
    # Apply date filters
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=end_date)
        except ValueError:
            pass
    
    # Apply status filter
    if status_filter and status_filter != 'all':
        orders = orders.filter(status=status_filter)
    
    # Create workbook with multiple sheets
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # =================================================================
    # SHEET 1: Orders Summary
    # =================================================================
    ws_summary = wb.create_sheet('Orders Summary')
    
    # Headers for summary sheet
    summary_headers = [
        'Order Number', 'Order Date', 'Customer Name', 'Customer Email', 
        'Status', 'Payment Status', 'Payment Method', 'Items Count',
        'Subtotal', 'Discount', 'Final Amount', 'Coupon Used',
        'Shipping City', 'Shipping State', 'Shipping Country', 'Created At'
    ]
    
    # Style headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2874F0', end_color='2874F0', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Auto-adjust column width
        column_letter = get_column_letter(col_num)
        ws_summary.column_dimensions[column_letter].width = max(len(header) + 2, 12)
    
    # Add summary data
    for row_num, order in enumerate(orders, 2):
        ws_summary.cell(row=row_num, column=1, value=order.order_number)
        ws_summary.cell(row=row_num, column=2, value=order.created_at.strftime('%Y-%m-%d'))
        ws_summary.cell(row=row_num, column=3, value=f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username)
        ws_summary.cell(row=row_num, column=4, value=order.user.email)
        ws_summary.cell(row=row_num, column=5, value=order.get_status_display())
        ws_summary.cell(row=row_num, column=6, value=order.payment_status.title())
        ws_summary.cell(row=row_num, column=7, value=order.payment_method.title())
        ws_summary.cell(row=row_num, column=8, value=order.items.count())
        ws_summary.cell(row=row_num, column=9, value=float(order.total_amount))
        ws_summary.cell(row=row_num, column=10, value=float(order.discount_amount))
        ws_summary.cell(row=row_num, column=11, value=float(order.final_amount))
        ws_summary.cell(row=row_num, column=12, value=order.coupon.code if order.coupon else 'None')
        ws_summary.cell(row=row_num, column=13, value=order.shipping_city)
        ws_summary.cell(row=row_num, column=14, value=order.shipping_state)
        ws_summary.cell(row=row_num, column=15, value=order.shipping_country)
        ws_summary.cell(row=row_num, column=16, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # =================================================================
    # SHEET 2: Order Items Detail
    # =================================================================
    ws_items = wb.create_sheet('Order Items Detail')
    
    # Headers for items sheet
    items_headers = [
        'Order Number', 'Order Date', 'Customer Name', 'Order Status',
        'Product Name', 'Product SKU', 'Size', 'Quantity', 'Unit Price', 
        'Total Price', 'Product Category', 'Product Brand'
    ]
    
    for col_num, header in enumerate(items_headers, 1):
        cell = ws_items.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        column_letter = get_column_letter(col_num)
        ws_items.column_dimensions[column_letter].width = max(len(header) + 2, 12)
    
    # Add items data
    item_row = 2
    for order in orders:
        for item in order.items.all():
            ws_items.cell(row=item_row, column=1, value=order.order_number)
            ws_items.cell(row=item_row, column=2, value=order.created_at.strftime('%Y-%m-%d'))
            ws_items.cell(row=item_row, column=3, value=f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username)
            ws_items.cell(row=item_row, column=4, value=order.get_status_display())
            ws_items.cell(row=item_row, column=5, value=item.product.name)
            ws_items.cell(row=item_row, column=6, value=getattr(item.product, 'sku', '') or f"PROD-{item.product.id}")
            ws_items.cell(row=item_row, column=7, value=item.size or 'N/A')
            ws_items.cell(row=item_row, column=8, value=item.quantity)
            ws_items.cell(row=item_row, column=9, value=float(item.price))
            ws_items.cell(row=item_row, column=10, value=float(item.get_total_price()))
            ws_items.cell(row=item_row, column=11, value=str(getattr(item.product, 'category', 'N/A')))
            ws_items.cell(row=item_row, column=12, value=getattr(item.product, 'brand', 'JOOG'))
            item_row += 1
    
    # =================================================================
    # SHEET 3: Customer Analysis
    # =================================================================
    ws_customers = wb.create_sheet('Customer Analysis')
    
    # Customer analysis headers
    customer_headers = [
        'Customer Name', 'Email', 'Total Orders', 'Total Spent', 
        'Average Order Value', 'First Order Date', 'Last Order Date',
        'Preferred Payment Method', 'Most Ordered City'
    ]
    
    for col_num, header in enumerate(customer_headers, 1):
        cell = ws_customers.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        column_letter = get_column_letter(col_num)
        ws_customers.column_dimensions[column_letter].width = max(len(header) + 2, 15)
    
    # Customer analysis data
    customer_stats = orders.values('user__email', 'user__first_name', 'user__last_name').annotate(
        total_orders=Count('id'),
        total_spent=Sum('final_amount'),
        avg_order_value=Avg('final_amount'),
        first_order=Min('created_at'),
        last_order=Max('created_at')
    ).order_by('-total_spent')
    
    cust_row = 2
    for customer in customer_stats:
        customer_orders = orders.filter(user__email=customer['user__email'])
        
        # Get most common payment method
        payment_methods = customer_orders.values('payment_method').annotate(
            count=Count('payment_method')
        ).order_by('-count').first()
        
        # Get most common shipping city
        cities = customer_orders.values('shipping_city').annotate(
            count=Count('shipping_city')
        ).order_by('-count').first()
        
        ws_customers.cell(row=cust_row, column=1, value=f"{customer['user__first_name'] or ''} {customer['user__last_name'] or ''}".strip() or 'N/A')
        ws_customers.cell(row=cust_row, column=2, value=customer['user__email'])
        ws_customers.cell(row=cust_row, column=3, value=customer['total_orders'])
        ws_customers.cell(row=cust_row, column=4, value=float(customer['total_spent'] or 0))
        ws_customers.cell(row=cust_row, column=5, value=float(customer['avg_order_value'] or 0))
        ws_customers.cell(row=cust_row, column=6, value=customer['first_order'].strftime('%Y-%m-%d') if customer['first_order'] else 'N/A')
        ws_customers.cell(row=cust_row, column=7, value=customer['last_order'].strftime('%Y-%m-%d') if customer['last_order'] else 'N/A')
        ws_customers.cell(row=cust_row, column=8, value=payment_methods['payment_method'].title() if payment_methods else 'N/A')
        ws_customers.cell(row=cust_row, column=9, value=cities['shipping_city'] if cities else 'N/A')
        cust_row += 1
    
    # =================================================================
    # SHEET 4: Sales Analytics
    # =================================================================
    ws_analytics = wb.create_sheet('Sales Analytics')
    
    # Analytics headers
    analytics_headers = [
        'Metric', 'Value', 'Period'
    ]
    
    for col_num, header in enumerate(analytics_headers, 1):
        cell = ws_analytics.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        column_letter = get_column_letter(col_num)
        ws_analytics.column_dimensions[column_letter].width = 20
    
    # Calculate analytics
    total_orders = orders.count()
    total_revenue = orders.aggregate(Sum('final_amount'))['final_amount__sum'] or 0
    total_discount = orders.aggregate(Sum('discount_amount'))['discount_amount__sum'] or 0
    avg_order_value = orders.aggregate(Avg('final_amount'))['final_amount__avg'] or 0
    
    # Status breakdown
    status_breakdown = orders.values('status').annotate(count=Count('id')).order_by('-count')
    
    # Payment method breakdown
    payment_breakdown = orders.values('payment_method').annotate(count=Count('id')).order_by('-count')
    
    # Monthly breakdown - SQLite compatible
    monthly_breakdown = orders.annotate(
        month=Extract('created_at', 'month'),
        year=Extract('created_at', 'year')
    ).values('month', 'year').annotate(
        order_count=Count('id'),
        revenue=Sum('final_amount')
    ).order_by('-year', '-month')[:12]
    
    # Add analytics data
    analytics_data = [
        ('Total Orders', total_orders, f"{start_date or 'All time'} to {end_date or 'Present'}"),
        ('Total Revenue', f"₹{total_revenue:,.2f}", f"{start_date or 'All time'} to {end_date or 'Present'}"),
        ('Total Discounts', f"₹{total_discount:,.2f}", f"{start_date or 'All time'} to {end_date or 'Present'}"),
        ('Average Order Value', f"₹{avg_order_value:,.2f}", f"{start_date or 'All time'} to {end_date or 'Present'}"),
        ('', '', ''),  # Empty row
        ('ORDER STATUS BREAKDOWN', '', ''),
    ]
    
    # Add status breakdown
    for status in status_breakdown:
        analytics_data.append((f"{status['status'].title()} Orders", status['count'], f"{status['count']/total_orders*100:.1f}% of total"))
    
    analytics_data.append(('', '', ''))  # Empty row
    analytics_data.append(('PAYMENT METHOD BREAKDOWN', '', ''))
    
    # Add payment breakdown
    for payment in payment_breakdown:
        analytics_data.append((f"{payment['payment_method'].title()} Payments", payment['count'], f"{payment['count']/total_orders*100:.1f}% of total"))
    
    analytics_data.append(('', '', ''))  # Empty row
    analytics_data.append(('MONTHLY BREAKDOWN', '', ''))
    
    # Add monthly breakdown
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for monthly in monthly_breakdown:
        month_name = months[int(monthly['month'])-1] if monthly['month'] else 'Unknown'
        analytics_data.append((f"{month_name} {monthly['year']}", f"{monthly['order_count']} orders, ₹{monthly['revenue']:,.2f}", 'Monthly Performance'))
    
    # Write analytics data
    for row_num, (metric, value, period) in enumerate(analytics_data, 2):
        ws_analytics.cell(row=row_num, column=1, value=metric)
        ws_analytics.cell(row=row_num, column=2, value=str(value))
        ws_analytics.cell(row=row_num, column=3, value=period)
        
        # Style section headers
        if 'BREAKDOWN' in metric:
            for col in range(1, 4):
                cell = ws_analytics.cell(row=row_num, column=col)
                cell.font = Font(bold=True, color='2874F0')
                cell.fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    date_suffix = f"_{start_date or 'all'}_{end_date or 'present'}"
    from django.utils import timezone
    now_local = timezone.localtime(timezone.now())
    filename = f'orders_detailed_report{date_suffix}_{now_local.strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@user_passes_test(is_staff_or_superuser)
def orders_report_dashboard(request):
    """Orders reporting dashboard with detailed analytics and export functionality"""
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status_filter = request.GET.get('status', 'all')
    
    orders = Order.objects.all().select_related('user', 'coupon').prefetch_related('items__product')
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=start_date)
        except ValueError:
            start_date = None
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=end_date)
        except ValueError:
            end_date = None
    
    if status_filter and status_filter != 'all':
        orders = orders.filter(status=status_filter)
    
    # Analytics calculations
    total_orders = orders.count()
    total_revenue = orders.aggregate(Sum('final_amount'))['final_amount__sum'] or 0
    total_discount = orders.aggregate(Sum('discount_amount'))['discount_amount__sum'] or 0
    avg_order_value = orders.aggregate(Avg('final_amount'))['final_amount__avg'] or 0
    
    # Status breakdown
    status_stats = orders.values('status').annotate(
        count=Count('id'),
        revenue=Sum('final_amount')
    ).order_by('-count')
    
    # Payment method breakdown
    payment_stats = orders.values('payment_method').annotate(
        count=Count('id'),
        revenue=Sum('final_amount')
    ).order_by('-count')
    
    # Top customers
    top_customers = orders.values(
        'user__email', 'user__first_name', 'user__last_name'
    ).annotate(
        total_orders=Count('id'),
        total_spent=Sum('final_amount'),
        avg_order_value=Avg('final_amount')
    ).order_by('-total_spent')[:10]
    
    # Monthly performance - SQLite compatible
    monthly_data = orders.annotate(
        month=Extract('created_at', 'month'),
        year=Extract('created_at', 'year')
    ).values('month', 'year').annotate(
        order_count=Count('id'),
        revenue=Sum('final_amount'),
        avg_order_value=Avg('final_amount')
    ).order_by('-year', '-month')[:12]
    
    # Recent orders for display
    recent_orders = orders.order_by('-created_at')[:10]
    
    # Top products
    from django.db.models import F
    top_products = OrderItem.objects.filter(
        order__in=orders
    ).values(
        'product__name', 'product__id'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('price')),
        order_count=Count('order', distinct=True)
    ).order_by('-total_revenue')[:10]
    
    # Geographic breakdown
    geographic_data = orders.values('shipping_state').annotate(
        order_count=Count('id'),
        revenue=Sum('final_amount')
    ).order_by('-order_count')[:10]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'status_filter': status_filter,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'total_discount': total_discount,
        'avg_order_value': avg_order_value,
        'status_stats': status_stats,
        'payment_stats': payment_stats,
        'top_customers': top_customers,
        'monthly_data': monthly_data,
        'recent_orders': recent_orders,
        'top_products': top_products,
        'geographic_data': geographic_data,
        'order_status_choices': Order.STATUS_CHOICES,
    }
    
    return render(request, 'invoices/orders_report_dashboard.html', context)


@user_passes_test(is_staff_or_superuser)
def export_sales_report_excel(request):
    """Export comprehensive sales report to Excel"""
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status_filter = request.GET.get('status', 'all')
    
    orders = Order.objects.all().select_related('user', 'coupon').prefetch_related('items__product')
    
    # Apply filters
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=end_date)
        except ValueError:
            pass
    
    if status_filter and status_filter != 'all':
        orders = orders.filter(status=status_filter)
    
    # Create workbook with multiple sheets
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # =================================================================
    # SHEET 1: Sales Summary Report
    # =================================================================
    ws_sales = wb.create_sheet('Sales Summary')
    
    # Style headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1f4e79', end_color='1f4e79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Sales summary headers
    sales_headers = [
        'Order Number', 'Date', 'Customer Name', 'Customer Email', 'Status',
        'Payment Method', 'Items Count', 'Subtotal', 'Discount', 'Final Amount',
        'Shipping City', 'Shipping State'
    ]
    
    for col_num, header in enumerate(sales_headers, 1):
        cell = ws_sales.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        column_letter = get_column_letter(col_num)
        ws_sales.column_dimensions[column_letter].width = max(len(header) + 2, 12)
    
    # Add sales data
    for row_num, order in enumerate(orders, 2):
        ws_sales.cell(row=row_num, column=1, value=order.order_number)
        ws_sales.cell(row=row_num, column=2, value=order.created_at.strftime('%Y-%m-%d'))
        ws_sales.cell(row=row_num, column=3, value=f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username)
        ws_sales.cell(row=row_num, column=4, value=order.user.email)
        ws_sales.cell(row=row_num, column=5, value=order.get_status_display())
        ws_sales.cell(row=row_num, column=6, value=order.payment_method.title())
        ws_sales.cell(row=row_num, column=7, value=order.items.count())
        ws_sales.cell(row=row_num, column=8, value=float(order.total_amount))
        ws_sales.cell(row=row_num, column=9, value=float(order.discount_amount))
        ws_sales.cell(row=row_num, column=10, value=float(order.final_amount))
        ws_sales.cell(row=row_num, column=11, value=order.shipping_city)
        ws_sales.cell(row=row_num, column=12, value=order.shipping_state)
    
    # =================================================================
    # SHEET 2: Product Sales Analysis
    # =================================================================
    ws_products = wb.create_sheet('Product Sales')
    
    # Product analysis headers
    product_headers = [
        'Product Name', 'Total Quantity Sold', 'Total Revenue', 'Orders Count',
        'Avg Price per Unit', 'Top Size', 'Category'
    ]
    
    for col_num, header in enumerate(product_headers, 1):
        cell = ws_products.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        column_letter = get_column_letter(col_num)
        ws_products.column_dimensions[column_letter].width = max(len(header) + 2, 15)
    
    # Get product sales data
    from django.db.models import F, Q, Count, Sum, Avg, Min, Max
    from orders.models import OrderItem
    product_sales = OrderItem.objects.filter(
        order__in=orders
    ).values(
        'product__name', 'product__category__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('price')),
        order_count=Count('order', distinct=True),
        avg_price=Avg('price')
    ).order_by('-total_revenue')
    
    # Add product data
    for row_num, product in enumerate(product_sales, 2):
        # Get most popular size for this product
        top_size = OrderItem.objects.filter(
            order__in=orders,
            product__name=product['product__name']
        ).values('size').annotate(
            size_count=Sum('quantity')
        ).order_by('-size_count').first()
        
        ws_products.cell(row=row_num, column=1, value=product['product__name'])
        ws_products.cell(row=row_num, column=2, value=product['total_quantity'])
        ws_products.cell(row=row_num, column=3, value=float(product['total_revenue']))
        ws_products.cell(row=row_num, column=4, value=product['order_count'])
        ws_products.cell(row=row_num, column=5, value=float(product['avg_price']))
        ws_products.cell(row=row_num, column=6, value=top_size['size'] if top_size else 'N/A')
        ws_products.cell(row=row_num, column=7, value=product['product__category__name'] or 'Uncategorized')
    
    # =================================================================
    # SHEET 3: Customer Sales Analysis
    # =================================================================
    ws_customers = wb.create_sheet('Customer Analysis')
    
    # Customer analysis headers
    customer_headers = [
        'Customer Name', 'Email', 'Total Orders', 'Total Spent', 'Avg Order Value',
        'First Purchase', 'Last Purchase', 'Favorite Product', 'Location'
    ]
    
    for col_num, header in enumerate(customer_headers, 1):
        cell = ws_customers.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        column_letter = get_column_letter(col_num)
        ws_customers.column_dimensions[column_letter].width = max(len(header) + 2, 15)
    
    # Get customer sales data
    customer_sales = orders.values(
        'user__email', 'user__first_name', 'user__last_name'
    ).annotate(
        total_orders=Count('id'),
        total_spent=Sum('final_amount'),
        avg_order_value=Avg('final_amount'),
        first_purchase=Min('created_at'),
        last_purchase=Max('created_at')
    ).order_by('-total_spent')
    
    # Add customer data
    for row_num, customer in enumerate(customer_sales, 2):
        customer_orders = orders.filter(user__email=customer['user__email'])
        
        # Get favorite product
        favorite_product = OrderItem.objects.filter(
            order__in=customer_orders
        ).values('product__name').annotate(
            quantity_sum=Sum('quantity')
        ).order_by('-quantity_sum').first()
        
        # Get most common location
        common_location = customer_orders.values('shipping_city').annotate(
            count=Count('shipping_city')
        ).order_by('-count').first()
        
        ws_customers.cell(row=row_num, column=1, value=f"{customer['user__first_name'] or ''} {customer['user__last_name'] or ''}".strip())
        ws_customers.cell(row=row_num, column=2, value=customer['user__email'])
        ws_customers.cell(row=row_num, column=3, value=customer['total_orders'])
        ws_customers.cell(row=row_num, column=4, value=float(customer['total_spent']))
        ws_customers.cell(row=row_num, column=5, value=float(customer['avg_order_value']))
        ws_customers.cell(row=row_num, column=6, value=customer['first_purchase'].strftime('%Y-%m-%d') if customer['first_purchase'] else 'N/A')
        ws_customers.cell(row=row_num, column=7, value=customer['last_purchase'].strftime('%Y-%m-%d') if customer['last_purchase'] else 'N/A')
        ws_customers.cell(row=row_num, column=8, value=favorite_product['product__name'] if favorite_product else 'N/A')
        ws_customers.cell(row=row_num, column=9, value=common_location['shipping_city'] if common_location else 'N/A')
    
    # =================================================================
    # SHEET 4: Sales Analytics Dashboard
    # =================================================================
    ws_dashboard = wb.create_sheet('Sales Dashboard')
    
    # Dashboard metrics - imports already available above
    total_orders = orders.count()
    total_revenue = orders.aggregate(Sum('final_amount'))['final_amount__sum'] or 0
    total_discount = orders.aggregate(Sum('discount_amount'))['discount_amount__sum'] or 0
    avg_order_value = orders.aggregate(Avg('final_amount'))['final_amount__avg'] or 0
    
    # Create dashboard layout
    dashboard_headers = ['Metric', 'Value', 'Details']
    
    for col_num, header in enumerate(dashboard_headers, 1):
        cell = ws_dashboard.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        column_letter = get_column_letter(col_num)
        ws_dashboard.column_dimensions[column_letter].width = 25
    
    # Dashboard data
    dashboard_data = [
        ('Total Orders', total_orders, f'Period: {start_date or "All time"} to {end_date or "Present"}'),
        ('Total Revenue', f'₹{total_revenue:,.2f}', f'Gross sales revenue'),
        ('Total Discounts', f'₹{total_discount:,.2f}', f'Total discount amount given'),
        ('Average Order Value', f'₹{avg_order_value:,.2f}', f'Average revenue per order'),
        ('', '', ''),  # Empty row
    ]
    
    # Status breakdown
    status_breakdown = orders.values('status').annotate(count=Count('id'), revenue=Sum('final_amount')).order_by('-count')
    dashboard_data.append(('ORDER STATUS ANALYSIS', '', ''))
    for status in status_breakdown:
        percentage = (status['count'] / total_orders * 100) if total_orders > 0 else 0
        dashboard_data.append((
            f"{status['status'].title()} Orders",
            f"{status['count']} orders",
            f"{percentage:.1f}% of total | ₹{status['revenue']:,.2f} revenue"
        ))
    
    dashboard_data.append(('', '', ''))  # Empty row
    
    # Payment method breakdown
    payment_breakdown = orders.values('payment_method').annotate(count=Count('id'), revenue=Sum('final_amount')).order_by('-count')
    dashboard_data.append(('PAYMENT METHOD ANALYSIS', '', ''))
    for payment in payment_breakdown:
        percentage = (payment['count'] / total_orders * 100) if total_orders > 0 else 0
        dashboard_data.append((
            f"{payment['payment_method'].title()}",
            f"{payment['count']} transactions",
            f"{percentage:.1f}% of total | ₹{payment['revenue']:,.2f} revenue"
        ))
    
    # Write dashboard data
    for row_num, (metric, value, details) in enumerate(dashboard_data, 2):
        ws_dashboard.cell(row=row_num, column=1, value=metric)
        ws_dashboard.cell(row=row_num, column=2, value=str(value))
        ws_dashboard.cell(row=row_num, column=3, value=details)
        
        # Style headers
        if 'ANALYSIS' in metric:
            for col in range(1, 4):
                cell = ws_dashboard.cell(row=row_num, column=col)
                cell.font = Font(bold=True, color='1f4e79')
                cell.fill = PatternFill(start_color='e7f3ff', end_color='e7f3ff', fill_type='solid')
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    date_suffix = f"_{start_date or 'all'}_{end_date or 'present'}"
    from django.utils import timezone
    now_local = timezone.localtime(timezone.now())
    filename = f'sales_report{date_suffix}_{now_local.strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@user_passes_test(is_staff_or_superuser)
def sales_reports_view(request):
    """JOOG-styled sales reports dashboard with comprehensive analytics"""
    from django.db.models import Count, Sum, Avg, F
    from datetime import datetime
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status_filter = request.GET.get('status', 'all')
    
    orders = Order.objects.all().select_related('user', 'coupon').prefetch_related('items__product')
    
    # Apply date filters
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=start_date)
        except ValueError:
            start_date = None
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=end_date)
        except ValueError:
            end_date = None
    
    # Apply status filter
    if status_filter and status_filter != 'all':
        orders = orders.filter(status=status_filter)
    
    # Calculate main statistics
    total_orders = orders.count()
    total_revenue = orders.aggregate(Sum('final_amount'))['final_amount__sum'] or 0
    total_discount = orders.aggregate(Sum('discount_amount'))['discount_amount__sum'] or 0
    avg_order_value = orders.aggregate(Avg('final_amount'))['final_amount__avg'] or 0
    
    # Status breakdown
    status_stats = orders.values('status').annotate(
        count=Count('id'),
        revenue=Sum('final_amount')
    ).order_by('-count')
    
    # Payment method breakdown
    payment_stats = orders.values('payment_method').annotate(
        count=Count('id'),
        revenue=Sum('final_amount')
    ).order_by('-count')
    
    # Top selling products
    from orders.models import OrderItem
    top_products = OrderItem.objects.filter(
        order__in=orders
    ).values(
        'product__name', 'product__category__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('price')),
        order_count=Count('order', distinct=True)
    ).order_by('-total_revenue')[:10]
    
    # Top customers
    top_customers = orders.values(
        'user__email', 'user__first_name', 'user__last_name'
    ).annotate(
        total_orders=Count('id'),
        total_spent=Sum('final_amount'),
        avg_order_value=Avg('final_amount')
    ).order_by('-total_spent')[:10]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'status_filter': status_filter,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'total_discount': total_discount,
        'avg_order_value': avg_order_value,
        'status_stats': status_stats,
        'payment_stats': payment_stats,
        'top_products': top_products,
        'top_customers': top_customers,
    }
    
    return render(request, 'invoices/sales_reports.html', context)
