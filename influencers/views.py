from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Sum, Count
from django.core.paginator import Paginator
from orders.models import Order
from .models import Influencer, InfluencerLoginLog
from .forms import InfluencerLoginForm

def info_view(request):
    """Information page about the influencer system"""
    # Get some test influencers to display
    test_influencers = Influencer.objects.filter(is_active=True)[:5]
    
    context = {
        'test_influencers': test_influencers,
    }
    
    return render(request, 'influencers/info.html', context)

def access_view(request):
    """Direct access page with credentials for testing"""
    return render(request, 'influencers/access_page.html')

def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def log_login_attempt(influencer, request, success=True):
    """Log influencer login attempt"""
    InfluencerLoginLog.objects.create(
        influencer=influencer,
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
        success=success
    )

def influencer_login_view(request):
    """Login page for influencers"""
    # Check if already authenticated via URL parameters
    username = request.GET.get('username')
    api_key = request.GET.get('api_key')
    
    if username and api_key:
        try:
            influencer = Influencer.objects.get(
                username=username, 
                api_key=api_key, 
                is_active=True
            )
            # Store in session
            request.session['influencer_id'] = influencer.id
            request.session['influencer_username'] = influencer.username
            
            # Update last login and log the attempt
            influencer.update_last_login()
            log_login_attempt(influencer, request, success=True)
            
            messages.success(request, f'Welcome back, {influencer.name}!')
            return redirect('influencers:dashboard')
        except Influencer.DoesNotExist:
            messages.error(request, 'Invalid credentials. Please check your login link.')
    
    # Handle form submission
    if request.method == 'POST':
        form = InfluencerLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            api_key = form.cleaned_data['api_key']
            
            try:
                influencer = Influencer.objects.get(
                    username=username, 
                    api_key=api_key, 
                    is_active=True
                )
                # Store in session
                request.session['influencer_id'] = influencer.id
                request.session['influencer_username'] = influencer.username
                
                # Update last login and log the attempt
                influencer.update_last_login()
                log_login_attempt(influencer, request, success=True)
                
                messages.success(request, f'Welcome back, {influencer.name}!')
                return redirect('influencers:dashboard')
            except Influencer.DoesNotExist:
                # Log failed attempt if we can find the influencer by username
                try:
                    influencer = Influencer.objects.get(username=username)
                    log_login_attempt(influencer, request, success=False)
                except Influencer.DoesNotExist:
                    pass
                
                messages.error(request, 'Invalid credentials. Please check your username and API key.')
    else:
        form = InfluencerLoginForm()
    
    return render(request, 'influencers/login.html', {'form': form})

def influencer_required(view_func):
    """Decorator to require influencer authentication"""
    def wrapper(request, *args, **kwargs):
        if 'influencer_id' not in request.session:
            messages.error(request, 'Please log in to access your dashboard.')
            return redirect('influencers:login')
        
        try:
            influencer = Influencer.objects.get(
                id=request.session['influencer_id'], 
                is_active=True
            )
            request.influencer = influencer
        except Influencer.DoesNotExist:
            messages.error(request, 'Your account is not active. Please contact support.')
            return redirect('influencers:login')
        
        return view_func(request, *args, **kwargs)
    return wrapper

@influencer_required
def dashboard_view(request):
    """Main dashboard for influencers"""
    influencer = request.influencer
    
    # Get assigned coupons
    assigned_coupons = influencer.coupons.all().select_related('coupon')
    coupon_codes = [ic.coupon.code for ic in assigned_coupons]
    
    # Get orders using influencer's coupons
    orders = Order.objects.filter(
        coupon__code__in=coupon_codes,
        payment_status='completed'
    ).select_related('user', 'coupon').prefetch_related('items__product').order_by('-created_at')
    
    # Calculate statistics
    total_orders = orders.count()
    total_revenue = orders.aggregate(total=Sum('final_amount'))['total'] or 0
    total_discount_given = orders.aggregate(total=Sum('discount_amount'))['total'] or 0
    commission_earned = (total_revenue * influencer.commission_rate) / 100
    
    # Recent orders (last 10)
    recent_orders = orders[:10]
    
    # Coupon performance
    coupon_stats = []
    for ic in assigned_coupons:
        coupon_orders = orders.filter(coupon=ic.coupon)
        coupon_revenue = coupon_orders.aggregate(total=Sum('final_amount'))['total'] or 0
        coupon_stats.append({
            'coupon': ic.coupon,
            'orders_count': coupon_orders.count(),
            'revenue': coupon_revenue,
            'usage_count': ic.coupon.used_count,
            'max_uses': ic.coupon.max_uses,
            'discount_type': ic.coupon.discount_type,
            'discount_value': ic.coupon.discount_value,
        })
    
    context = {
        'influencer': influencer,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'total_discount_given': total_discount_given,
        'commission_earned': commission_earned,
        'recent_orders': recent_orders,
        'coupon_stats': coupon_stats,
        'assigned_coupons_count': len(coupon_codes),
    }
    
    return render(request, 'influencers/dashboard.html', context)

@influencer_required
def orders_view(request):
    """View all orders for the influencer"""
    influencer = request.influencer
    
    # Get assigned coupons
    assigned_coupons = influencer.coupons.all().select_related('coupon')
    coupon_codes = [ic.coupon.code for ic in assigned_coupons]
    
    # Get orders using influencer's coupons
    orders = Order.objects.filter(
        coupon__code__in=coupon_codes
    ).select_related('user', 'coupon').prefetch_related('items__product').order_by('-created_at')
    
    # Filter by status if specified
    status_filter = request.GET.get('status')
    if status_filter and status_filter != 'all':
        orders = orders.filter(status=status_filter)
    
    # Filter by coupon if specified
    coupon_filter = request.GET.get('coupon')
    if coupon_filter and coupon_filter != 'all':
        orders = orders.filter(coupon__code=coupon_filter)
    
    # Pagination
    paginator = Paginator(orders, 20)
    page_number = request.GET.get('page')
    orders_page = paginator.get_page(page_number)
    
    # Get filter options
    status_choices = Order.STATUS_CHOICES
    
    context = {
        'influencer': influencer,
        'orders': orders_page,
        'coupon_codes': coupon_codes,
        'status_choices': status_choices,
        'current_status': status_filter,
        'current_coupon': coupon_filter,
    }
    
    return render(request, 'influencers/orders.html', context)

@influencer_required
def profile_view(request):
    """Profile page for influencers"""
    influencer = request.influencer
    
    # Get assigned coupons
    assigned_coupons = influencer.coupons.all().select_related('coupon')
    coupon_codes = [ic.coupon for ic in assigned_coupons]
    
    # Get orders using influencer's coupons
    orders = Order.objects.filter(
        coupon__code__in=[ic.coupon.code for ic in assigned_coupons],
        payment_status='completed'
    ).select_related('user', 'coupon')
    
    # Calculate statistics
    total_orders = orders.count()
    total_revenue = orders.aggregate(total=Sum('final_amount'))['total'] or 0
    commission_earned = (total_revenue * influencer.commission_rate) / 100
    active_coupons = len([c for c in coupon_codes if c.active])
    
    stats = {
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'total_commission': commission_earned,
        'active_coupons': active_coupons,
    }
    
    context = {
        'influencer': influencer,
        'coupon_codes': coupon_codes,
        'stats': stats,
    }
    
    return render(request, 'influencers/profile.html', context)

@influencer_required
def monthly_report_view(request):
    """Monthly orders report for influencers"""
    from django.db.models import Sum, Count, Avg
    from datetime import datetime, date
    from django.utils import timezone
    import calendar
    
    influencer = request.influencer
    
    # Get current date
    now = timezone.now()
    current_year = now.year
    current_month = now.month
    
    # Get filters from request
    selected_year = int(request.GET.get('year', current_year))
    selected_month = int(request.GET.get('month', current_month))
    
    # Get assigned coupons
    assigned_coupons = influencer.coupons.all().select_related('coupon')
    coupon_codes = [ic.coupon.code for ic in assigned_coupons]
    
    # Filter orders by selected month and year
    orders = Order.objects.filter(
        coupon__code__in=coupon_codes,
        created_at__year=selected_year,
        created_at__month=selected_month
    ).select_related('user', 'coupon').prefetch_related('items__product').order_by('-created_at')
    
    # Calculate statistics for the month
    completed_orders = orders.filter(payment_status='completed')
    total_orders = orders.count()
    total_completed = completed_orders.count()
    total_revenue = completed_orders.aggregate(total=Sum('final_amount'))['total'] or 0
    total_discount_given = completed_orders.aggregate(total=Sum('discount_amount'))['total'] or 0
    commission_earned = (total_revenue * influencer.commission_rate) / 100
    avg_order_value = completed_orders.aggregate(avg=Avg('final_amount'))['avg'] or 0
    
    # Daily breakdown for the month
    import calendar
    from collections import defaultdict
    
    daily_stats = defaultdict(lambda: {'orders': 0, 'revenue': 0, 'commission': 0})
    
    for order in completed_orders:
        day = order.created_at.day
        daily_stats[day]['orders'] += 1
        daily_stats[day]['revenue'] += float(order.final_amount)
        daily_stats[day]['commission'] += float(order.final_amount * influencer.commission_rate / 100)
    
    # Convert to list for template
    days_in_month = calendar.monthrange(selected_year, selected_month)[1]
    daily_breakdown = []
    for day in range(1, days_in_month + 1):
        daily_breakdown.append({
            'day': day,
            'orders': daily_stats[day]['orders'],
            'revenue': daily_stats[day]['revenue'],
            'commission': daily_stats[day]['commission']
        })
    
    # Coupon performance for the month
    coupon_stats = []
    for ic in assigned_coupons:
        coupon_orders = orders.filter(coupon=ic.coupon, payment_status='completed')
        coupon_revenue = coupon_orders.aggregate(total=Sum('final_amount'))['total'] or 0
        coupon_stats.append({
            'coupon': ic.coupon,
            'orders_count': coupon_orders.count(),
            'revenue': coupon_revenue,
            'commission': (coupon_revenue * influencer.commission_rate) / 100,
        })
    
    # Generate year and month options for filters
    years = list(range(current_year - 2, current_year + 1))
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    
    context = {
        'influencer': influencer,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'selected_month_name': calendar.month_name[selected_month],
        'years': years,
        'months': months,
        'total_orders': total_orders,
        'total_completed': total_completed,
        'total_revenue': total_revenue,
        'total_discount_given': total_discount_given,
        'commission_earned': commission_earned,
        'avg_order_value': avg_order_value,
        'orders': orders[:50],  # Show first 50 orders
        'daily_breakdown': daily_breakdown,
        'coupon_stats': coupon_stats,
        'total_orders_count': total_orders,
    }
    
    return render(request, 'influencers/monthly_report.html', context)

@influencer_required
def export_monthly_report_excel(request):
    """Export monthly orders report to Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.db.models import Sum, Count, Avg
    from datetime import datetime
    from django.utils import timezone
    import calendar
    
    influencer = request.influencer
    
    # Get filters from request
    now = timezone.now()
    selected_year = int(request.GET.get('year', now.year))
    selected_month = int(request.GET.get('month', now.month))
    
    # Get assigned coupons
    assigned_coupons = influencer.coupons.all().select_related('coupon')
    coupon_codes = [ic.coupon.code for ic in assigned_coupons]
    
    # Filter orders by selected month and year
    orders = Order.objects.filter(
        coupon__code__in=coupon_codes,
        created_at__year=selected_year,
        created_at__month=selected_month
    ).select_related('user', 'coupon').prefetch_related('items__product').order_by('-created_at')
    
    completed_orders = orders.filter(payment_status='completed')
    
    # Create workbook and worksheets
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Summary Sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Header styles
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=12)
    
    # Summary sheet content
    month_name = calendar.month_name[selected_month]
    summary_ws['A1'] = f"Monthly Report - {month_name} {selected_year}"
    summary_ws['A1'].font = header_font
    summary_ws.merge_cells('A1:D1')
    
    summary_ws['A3'] = "Influencer:"
    summary_ws['B3'] = influencer.name
    summary_ws['A4'] = "Username:"
    summary_ws['B4'] = influencer.username
    summary_ws['A5'] = "Commission Rate:"
    summary_ws['B5'] = f"{influencer.commission_rate}%"
    
    # Statistics
    total_orders = orders.count()
    total_completed = completed_orders.count()
    total_revenue = completed_orders.aggregate(total=Sum('final_amount'))['total'] or 0
    total_discount = completed_orders.aggregate(total=Sum('discount_amount'))['total'] or 0
    commission_earned = (total_revenue * influencer.commission_rate) / 100
    avg_order_value = completed_orders.aggregate(avg=Avg('final_amount'))['avg'] or 0
    
    summary_ws['A7'] = "MONTHLY STATISTICS"
    summary_ws['A7'].font = subheader_font
    
    summary_ws['A8'] = "Total Orders:"
    summary_ws['B8'] = total_orders
    summary_ws['A9'] = "Completed Orders:"
    summary_ws['B9'] = total_completed
    summary_ws['A10'] = "Total Revenue:"
    summary_ws['B10'] = f"₹{total_revenue:.2f}"
    summary_ws['A11'] = "Commission Earned:"
    summary_ws['B11'] = f"₹{commission_earned:.2f}"
    summary_ws['A12'] = "Average Order Value:"
    summary_ws['B12'] = f"₹{avg_order_value:.2f}"
    summary_ws['A13'] = "Total Discount Given:"
    summary_ws['B13'] = f"₹{total_discount:.2f}"
    
    # Orders Detail Sheet
    orders_ws = wb.create_sheet("Orders Detail")
    
    # Headers for orders
    headers = ['Order Number', 'Date', 'Customer', 'Status', 'Payment Status', 
              'Total Amount', 'Discount', 'Final Amount', 'Coupon Used', 'Commission']
    
    for col, header in enumerate(headers, 1):
        cell = orders_ws.cell(row=1, column=col, value=header)
        cell.font = subheader_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add order data
    for row, order in enumerate(orders, 2):
        commission = (float(order.final_amount) * float(influencer.commission_rate)) / 100 if order.payment_status == 'completed' else 0
        
        orders_ws.cell(row=row, column=1, value=order.order_number)
        orders_ws.cell(row=row, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
        orders_ws.cell(row=row, column=3, value=order.user.username if order.user else 'N/A')
        orders_ws.cell(row=row, column=4, value=order.get_status_display())
        orders_ws.cell(row=row, column=5, value=order.payment_status.title())
        orders_ws.cell(row=row, column=6, value=float(order.total_amount))
        orders_ws.cell(row=row, column=7, value=float(order.discount_amount))
        orders_ws.cell(row=row, column=8, value=float(order.final_amount))
        orders_ws.cell(row=row, column=9, value=order.coupon.code if order.coupon else 'N/A')
        orders_ws.cell(row=row, column=10, value=commission)
    
    # Coupon Performance Sheet
    coupon_ws = wb.create_sheet("Coupon Performance")
    
    coupon_headers = ['Coupon Code', 'Orders Count', 'Revenue Generated', 'Commission Earned', 'Discount Type', 'Discount Value']
    
    for col, header in enumerate(coupon_headers, 1):
        cell = coupon_ws.cell(row=1, column=col, value=header)
        cell.font = subheader_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add coupon data
    for row, ic in enumerate(assigned_coupons, 2):
        coupon_orders = completed_orders.filter(coupon=ic.coupon)
        coupon_revenue = coupon_orders.aggregate(total=Sum('final_amount'))['total'] or 0
        coupon_commission = (coupon_revenue * influencer.commission_rate) / 100
        
        coupon_ws.cell(row=row, column=1, value=ic.coupon.code)
        coupon_ws.cell(row=row, column=2, value=coupon_orders.count())
        coupon_ws.cell(row=row, column=3, value=float(coupon_revenue))
        coupon_ws.cell(row=row, column=4, value=float(coupon_commission))
        coupon_ws.cell(row=row, column=5, value=ic.coupon.get_discount_type_display())
        coupon_ws.cell(row=row, column=6, value=float(ic.coupon.discount_value))
    
    # Adjust column widths
    for ws in [summary_ws, orders_ws, coupon_ws]:
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # Find the first non-merged cell to get column letter
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            # Skip if no valid column letter found
            if not column_letter:
                continue
                
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"monthly_report_{influencer.username}_{month_name}_{selected_year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def logout_view(request):
    """Logout influencer"""
    if 'influencer_id' in request.session:
        del request.session['influencer_id']
    if 'influencer_username' in request.session:
        del request.session['influencer_username']
    
    messages.success(request, 'You have been logged out successfully.')
    return redirect('influencers:login')
